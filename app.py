from flask import Flask, request
from flask import json
from openpyxl import load_workbook
import os
import textwrap

app = Flask(__name__, static_url_path='')


def organiza_dados(seqman):
    seqIndice = []
    sequencia = []
    cont = enumerate(seqman)
    for i, item in enumerate(seqman):
        wrapper = textwrap.TextWrapper(width=80)
        lista = wrapper.wrap(text=item)
        # del seqman[i]
        # print(lista)
        sequencia.extend(lista)
        seqIndice.append(i)
        j = 1
        while (j < len(lista)):
            seqIndice.append('XXX')
            j += 1
        """ #seqindice[i] = i+1
        j = 1
        while (j < len(lista)):
            i += 1
            j += 1
            # seqindice[i] = 'XXX'
            cont += 1 """
    print(seqIndice, sequencia)
    return seqIndice, sequencia


def cria_pagina(event_data, page, cont, tipo, seqIndice):
    if(page > 0):
        # Se não for a primeira página deleta os 32 primeiros itens, pois os mesmo já foram escritos
        del event_data[0:32]
        del seqIndice[0:32]
    # load excel file
    workbook = load_workbook(filename=os.path.join(
        'static', "SMout.xlsx"))  # Carrega o arquivo de saída
    page += 1   # Incrementa número da página (sheet)
    # insert at the end (default)
    name = tipo + str(page)  # Obtém o NOME do sheet, que fica escrito na aba
    sheet = workbook[name]
    # Atribui o item de manobra célula por célula
    for i, new_val in enumerate(event_data):
        num = i+8
        cell_ix = "A%d" % num
        cell_time = "B%d" % num
        cell_im = "C%d" % num
        sheet[cell_im] = new_val
        sheet[cell_ix] = seqIndice[i]
        sheet[cell_time] = ':'
        if(seqIndice[i] == 'XXX'):
            sheet[cell_time] = 'XXX'
            print('Entrooooooooooooooooooooooooou')
        cont += 1
        if (i >= 31):  # Caso a SM tenha mais de 32 itens, chama-se a função de maneira recursiva para escrever na próxima página
            workbook.save(filename=os.path.join('static', "SMout.xlsx"))
            page = cria_pagina(event_data, page, cont, tipo, seqIndice)
            return page
        # Caso se chega ao fim do Array da SM, atribui-se os XXXXX a próxima linha CASO o espaço não tenha acabado
        elif(i == len(event_data)-1):
            if((i+1) % 32 != 0):
                num += 1
                cell_ix = "A%d" % num
                cell_time = "B%d" % num
                cell_im = "C%d" % num
                sheet[cell_im] = "XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX"
                sheet[cell_ix] = "XXX"
                sheet[cell_time] = "XXX"
            for x in range(page+1, 11):  # Remove todos os sheets excedentes
                workbook.remove(workbook[tipo + str(x)])
            workbook.save(filename=os.path.join('static', "SMout.xlsx"))
            return page
    # Caso o Array esteja vazio, deleta-se os sheets restantes e salva o arquivo
    for x in range(page, 11):
        workbook.remove(workbook[tipo + str(x)])
    workbook.save(filename=os.path.join('static', "SMout.xlsx"))


# Recebe-se o método POST com o objeto JSON que contém os arrays
@app.route("/", methods=['POST'])
def index():
    event_data = request.json   # Grava os dados JSON na variável
    # cria e salva um novo arquivo Excel
    workbook = load_workbook(filename=os.path.join(
        'static', "SM.xlsx"))  # Carrega o template original
    # Salva em um arquivo de saída
    workbook.save(filename=os.path.join('static', "SMout.xlsx"))
    # workbook = load_workbook(filename=os.path.join('static', "SMout.xlsx"))
    seqIndice, seqIsolar = organiza_dados(event_data['seqIsolar'])
    print('tamanho inidice' + str(len(seqIndice)) +
          'tamanho seq ' + str(len(seqIsolar)))
    # Escreve nos sheets de ISOLAR
    cria_pagina(seqIsolar, 0, 1, "ISOLAR-", seqIndice)
    print(event_data['seqNormalizar'])
    seqIndice, seqNormalizar = organiza_dados(event_data['seqNormalizar'])
    # Escreve nos sheets de normalizar
    cria_pagina(seqNormalizar, 0, 1, "NORMALIZAR-", seqIndice)

    file_name = 'SMout.xlsx'
    return app.send_static_file(filename=file_name)


# Caso haja um arquivo de saída, deleta-o
if os.path.exists(os.path.join('static', "SMout.xlsx")):
    os.remove(os.path.join('static', "SMout.xlsx"))
else:
    print("Extra file doesn't exists")


@ app.after_request
def set_headers(response):
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Headers"] = "*"
    response.headers["Access-Control-Allow-Methods"] = "*"
    return response


if __name__ == "__main__":
    app.run(host='127.0.0.1', port=5006)
